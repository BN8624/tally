# 회계프로그램 엑셀에서 실제 상세 거래만 엄격하게 읽습니다.
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re
from typing import BinaryIO

import pandas as pd
from openpyxl import load_workbook


class InputWorkbookError(ValueError):
    """입력 엑셀의 구조나 값이 요구 형식과 다를 때 발생합니다."""


REQUIRED_COLUMNS = (
    "구분",
    "전표일자",
    "거래처",
    "품명",
    "공급가액",
    "부가세",
    "합계",
    "매입/매출 유형",
    "계정코드",
    "계정과목",
    "카드사명",
    "카드번호",
)

HEADER_ALIASES = {
    "구분": {"구분"},
    "전표일자": {"전표일자"},
    "거래처": {"거래처"},
    "품명": {"품명"},
    "공급가액": {"공급가액"},
    "부가세": {"부가세", "세액"},
    "합계": {"합계", "합계금액"},
    "매입/매출 유형": {"매입/매출유형", "매입매출유형"},
    "계정과목": {"계정과목"},
    "카드사명": {"카드사명"},
    "카드번호": {"카드번호"},
}

OUTPUT_COLUMNS = (
    "row_id",
    "sheet",
    "source_row",
    "division",
    "date",
    "month",
    "vendor",
    "item",
    "supply_amount",
    "tax_amount",
    "total_amount",
    "original_type",
    "account_code",
    "account_name",
    "card_company",
    "card_number",
)


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).strip()


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _clean_code(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text[:-2] if text.endswith(".0") and text[:-2].isdigit() else text


def _parse_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not isinstance(value, str):
        return None
    text = value.strip()
    for pattern in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def _parse_amount(value: object, *, sheet: str, row: int, column: str) -> Decimal:
    if isinstance(value, bool) or value is None:
        raise InputWorkbookError(
            f"금액 형식 오류. 시트={sheet}, 행={row}, 열={column}, 값={value!r}. "
            "숫자 또는 천 단위 구분기호가 있는 숫자인지 확인하세요."
        )
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip().replace(",", "")
    if not text:
        raise InputWorkbookError(
            f"금액 형식 오류. 시트={sheet}, 행={row}, 열={column}, 값은 비어 있습니다. "
            "원본 상세 거래의 금액을 확인하세요."
        )
    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1]}"
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise InputWorkbookError(
            f"금액 형식 오류. 시트={sheet}, 행={row}, 열={column}, 값={value!r}. "
            "임의로 추정하지 않았습니다. 원본 값을 확인하세요."
        ) from exc


def _normalize_type(value: object) -> str:
    text = _clean_text(value)
    if "." in text:
        text = text.split(".", 1)[1].strip()
    return text


def _column_map(headers: tuple[object, ...]) -> tuple[dict[str, int], list[str]]:
    normalized = [_normalize_header(value) for value in headers]
    mapping: dict[str, int] = {}
    recognized: list[str] = []
    for logical, aliases in HEADER_ALIASES.items():
        for index, header in enumerate(normalized):
            if header in aliases:
                mapping[logical] = index
                recognized.append(logical)
                break

    account_index = mapping.get("계정과목")
    if account_index is not None:
        preceding_codes = [
            index for index, header in enumerate(normalized[:account_index]) if header.casefold() == "code"
        ]
        if preceding_codes:
            mapping["계정코드"] = preceding_codes[-1]
            recognized.append("계정코드")
    return mapping, recognized


def _find_header(workbook) -> tuple[object, int, dict[str, int]]:
    best_sheet = "알 수 없음"
    best_recognized: list[str] = []
    for worksheet in workbook.worksheets:
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=min(30, worksheet.max_row), values_only=True),
            start=1,
        ):
            mapping, recognized = _column_map(row)
            if len(recognized) > len(best_recognized):
                best_sheet = worksheet.title
                best_recognized = recognized
            if all(column in mapping for column in REQUIRED_COLUMNS):
                return worksheet, row_number, mapping

    missing = [column for column in REQUIRED_COLUMNS if column not in best_recognized]
    raise InputWorkbookError(
        "필수 열을 찾지 못했습니다. "
        f"찾지 못한 열={', '.join(missing) or '없음'}, "
        f"인식한 열={', '.join(best_recognized) or '없음'}, "
        f"문제가 발생한 시트={best_sheet}. "
        "전체 매입매출장 원본 형식과 열 제목을 확인하세요."
    )


def parse_workbook(source: str | Path | BinaryIO) -> pd.DataFrame:
    """전체 매입매출장에서 날짜가 있는 상세 거래만 반환합니다."""
    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as exc:
        raise InputWorkbookError(f"엑셀 파일을 열 수 없습니다. 파일 형식과 암호 설정을 확인하세요. {exc}") from exc

    worksheet, header_row, columns = _find_header(workbook)
    transactions: list[dict[str, object]] = []
    for source_row, row in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        transaction_date = _parse_date(row[columns["전표일자"]])
        if transaction_date is None:
            continue
        division = _clean_text(row[columns["구분"]])
        if division not in {"매입", "매출"}:
            raise InputWorkbookError(
                f"구분 값 오류. 시트={worksheet.title}, 행={source_row}, 값={division!r}. "
                "매입 또는 매출인지 확인하세요."
            )
        transactions.append(
            {
                "row_id": f"{worksheet.title}:{source_row}",
                "sheet": worksheet.title,
                "source_row": source_row,
                "division": division,
                "date": transaction_date,
                "month": transaction_date.strftime("%Y-%m"),
                "vendor": _clean_text(row[columns["거래처"]]),
                "item": _clean_text(row[columns["품명"]]),
                "supply_amount": _parse_amount(
                    row[columns["공급가액"]], sheet=worksheet.title, row=source_row, column="공급가액"
                ),
                "tax_amount": _parse_amount(
                    row[columns["부가세"]], sheet=worksheet.title, row=source_row, column="부가세"
                ),
                "total_amount": _parse_amount(
                    row[columns["합계"]], sheet=worksheet.title, row=source_row, column="합계"
                ),
                "original_type": _normalize_type(row[columns["매입/매출 유형"]]),
                "account_code": _clean_code(row[columns["계정코드"]]),
                "account_name": _clean_text(row[columns["계정과목"]]),
                "card_company": _clean_text(row[columns["카드사명"]]),
                "card_number": _clean_text(row[columns["카드번호"]]),
            }
        )
    return pd.DataFrame(transactions, columns=OUTPUT_COLUMNS)

