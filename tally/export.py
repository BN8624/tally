# 처리 결과를 인쇄 가능한 집계표와 검토·상세·검산 시트로 출력합니다.
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .core import ProcessingResult
from .settings import CompanySettings


NAVY = "1F4E78"
RED = "F4CCCC"
WHITE = "FFFFFF"
THIN_GRAY = Side(style="thin", color="B7B7B7")
LEDGER_INK = "404040"
LEDGER_RULE = Side(style="dotted", color="B7B7B7")
LEDGER_TOTAL_RULE = Side(style="medium", color="666666")
DETAIL_START_COLUMNS = (1, 5)
SHARED_VALUE_START_COLUMNS = (2, 5, 8, 11)
SUMMARY_END_COLUMN = 13
COUNT_FORMAT = '"("0")"'
MONEY_FORMAT = "#,##0;[Red](#,##0);-"
LEDGER_ROW_HEIGHT = 25


def _excel_value(value: object) -> object:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, Decimal):
        return int(value) if value == value.to_integral_value() else float(value)
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return value


def _period_label(months: list[str]) -> str:
    if not months:
        return "집계 기간 없음"
    years = {month[:4] for month in months}
    month_numbers = {int(month[5:]) for month in months}
    if len(years) == 1:
        year = months[0][:4]
        periods = (
            ({1, 2, 3}, "1기 예정"),
            ({4, 5, 6}, "1기 확정"),
            ({7, 8, 9}, "2기 예정"),
            ({10, 11, 12}, "2기 확정"),
        )
        for period_months, label in periods:
            if month_numbers and month_numbers.issubset(period_months):
                return f"{year}년 {label}"
    return f"{months[0]} ~ {months[-1]}"


def _set_summary_heading(sheet, company_name: str, months: list[str]) -> None:
    headings = (
        (1, 4, _period_label(months), "left", 10),
        (12, 13, "부가세 집계표", "right", 10),
    )
    for start_column, end_column, text, alignment, size in headings:
        sheet.merge_cells(
            start_row=1,
            start_column=start_column,
            end_row=1,
            end_column=end_column,
        )
        cell = sheet.cell(1, start_column, text)
        cell.font = Font(name="맑은 고딕", size=size, bold=True, color=LEDGER_INK)
        cell.alignment = Alignment(horizontal=alignment, vertical="center")


def _set_title(sheet, text: str, end_column: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = sheet.cell(1, 1, text)
    cell.font = Font(name="맑은 고딕", size=16, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 30


def _lookup(frame: pd.DataFrame, key_column: str, key: str, month: str, value: str) -> object:
    rows = frame[(frame[key_column].eq(key)) & (frame["month"].eq(month))]
    if rows.empty:
        return 0
    return _excel_value(rows.iloc[0][value])


def _month_label(month: str, months: list[str]) -> str:
    years = {value[:4] for value in months}
    if len(years) > 1:
        return f"{month[:4]}.{int(month[5:])}월"
    return f"{int(month[5:])}월"


def _has_values(
    frame: pd.DataFrame,
    key_column: str,
    key: str,
    fields: tuple[str, ...] = ("count", "supply_amount", "tax_amount", "total_amount"),
) -> bool:
    rows = frame[frame[key_column].eq(key)]
    return any(
        field in rows and pd.to_numeric(rows[field], errors="coerce").fillna(0).ne(0).any()
        for field in fields
    )


def _total_value(frame: pd.DataFrame, key_column: str, key: str, field: str) -> object:
    rows = frame[frame[key_column].eq(key)]
    if rows.empty or field not in rows:
        return 0
    return _excel_value(pd.to_numeric(rows[field], errors="coerce").fillna(0).sum())


def _style_ledger_cell(cell, *, bold: bool = False, total: bool = False) -> None:
    cell.fill = PatternFill("solid", fgColor=WHITE)
    cell.font = Font(name="맑은 고딕", size=9, bold=bold, color=LEDGER_INK)
    cell.alignment = Alignment(vertical="center")
    cell.border = Border(
        top=LEDGER_TOTAL_RULE if total else None,
        bottom=LEDGER_RULE,
    )


def _write_month_table(
    sheet,
    title_row: int,
    start_column: int,
    title: str,
    frame: pd.DataFrame,
    key_column: str,
    key: str,
    months: list[str],
    *,
    marker: str = "",
    exempt: bool = False,
    title_start_column: int | None = None,
) -> int:
    title_start = title_start_column or start_column
    end_column = start_column + 3
    sheet.merge_cells(
        start_row=title_row,
        start_column=title_start,
        end_row=title_row,
        end_column=end_column,
    )
    title_cell = sheet.cell(title_row, title_start, f"{marker}  {title}" if marker else title)
    _style_ledger_cell(title_cell, bold=True)
    title_cell.font = Font(name="맑은 고딕", size=10, bold=True, color=LEDGER_INK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    first_data_row = title_row + 1
    fields: tuple[str | None, ...] = (
        "count",
        "supply_amount",
        None if exempt else "tax_amount",
    )
    for month_index, month in enumerate(months):
        row = first_data_row + month_index
        month_cell = sheet.cell(row, start_column, _month_label(month, months))
        _style_ledger_cell(month_cell)
        month_cell.alignment = Alignment(horizontal="center", vertical="center")
        for offset, field in enumerate(fields, start=1):
            cell = sheet.cell(row, start_column + offset)
            if field is not None:
                cell.value = _lookup(frame, key_column, key, month, field)
            _style_ledger_cell(cell)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = COUNT_FORMAT if field == "count" else MONEY_FORMAT

    total_row = first_data_row + len(months)
    total_label = sheet.cell(total_row, start_column, "계")
    _style_ledger_cell(total_label, bold=True, total=True)
    total_label.alignment = Alignment(horizontal="center", vertical="center")
    for offset, field in enumerate(fields, start=1):
        column = start_column + offset
        cell = sheet.cell(total_row, column)
        if field is not None:
            letter = get_column_letter(column)
            cell.value = f"=SUM({letter}{first_data_row}:{letter}{total_row - 1})" if months else 0
        _style_ledger_cell(cell, bold=True, total=True)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = COUNT_FORMAT if field == "count" else MONEY_FORMAT
    return total_row


def _write_compact_month_table(
    sheet,
    title_row: int,
    start_column: int,
    title: str,
    frame: pd.DataFrame,
    key_column: str,
    key: str,
    months: list[str],
) -> int:
    sheet.merge_cells(
        start_row=title_row,
        start_column=start_column,
        end_row=title_row,
        end_column=start_column + 2,
    )
    title_cell = sheet.cell(title_row, start_column, title)
    _style_ledger_cell(title_cell, bold=True)
    title_cell.font = Font(name="맑은 고딕", size=10, bold=True, color=LEDGER_INK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    first_data_row = title_row + 1
    fields = ("count", "supply_amount", "tax_amount")
    for month_index, month in enumerate(months):
        row = first_data_row + month_index
        for offset, field in enumerate(fields):
            cell = sheet.cell(row, start_column + offset, _lookup(frame, key_column, key, month, field))
            _style_ledger_cell(cell)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = COUNT_FORMAT if field == "count" else MONEY_FORMAT

    total_row = first_data_row + len(months)
    for offset, field in enumerate(fields):
        column = start_column + offset
        letter = get_column_letter(column)
        cell = sheet.cell(
            total_row,
            column,
            f"=SUM({letter}{first_data_row}:{letter}{total_row - 1})" if months else 0,
        )
        _style_ledger_cell(cell, bold=True, total=True)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = COUNT_FORMAT if field == "count" else MONEY_FORMAT
    return total_row


def _write_total_amount_table(
    sheet,
    title_row: int,
    items: list[tuple[str, str]],
    frame: pd.DataFrame,
    months: list[str],
) -> tuple[int, list[int]]:
    for offset, (title, _key) in enumerate(items, start=2):
        cell = sheet.cell(title_row, offset, title)
        _style_ledger_cell(cell, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    first_data_row = title_row + 1
    for month_index, month in enumerate(months):
        row = first_data_row + month_index
        month_cell = sheet.cell(row, 1, _month_label(month, months))
        _style_ledger_cell(month_cell)
        month_cell.alignment = Alignment(horizontal="center", vertical="bottom")
        for column, (_title, key) in enumerate(items, start=2):
            cell = sheet.cell(row, column, _lookup(frame, "item", key, month, "total_amount"))
            _style_ledger_cell(cell)
            cell.alignment = Alignment(horizontal="right", vertical="bottom")
            cell.number_format = MONEY_FORMAT

    total_row = first_data_row + len(months)
    label = sheet.cell(total_row, 1, "계")
    _style_ledger_cell(label, bold=True, total=True)
    label.alignment = Alignment(horizontal="center", vertical="bottom")
    total_columns: list[int] = []
    for column in range(2, 2 + len(items)):
        letter = get_column_letter(column)
        cell = sheet.cell(
            total_row,
            column,
            f"=SUM({letter}{first_data_row}:{letter}{total_row - 1})" if months else 0,
        )
        _style_ledger_cell(cell, bold=True, total=True)
        cell.alignment = Alignment(horizontal="right", vertical="bottom")
        cell.number_format = MONEY_FORMAT
        total_columns.append(column)
    return total_row, total_columns


def _ordered_categories(settings: CompanySettings) -> list[str]:
    return [
        settings.account_146_label,
        "담배",
        "원재료(도급)",
        "제조경비",
        "도급경비",
        "기타",
    ]


def _write_shared_month_band(
    sheet,
    title_row: int,
    items: list[tuple[str, pd.DataFrame, str, str, bool, bool] | None],
    months: list[str],
    *,
    marker: str = "",
    first_title_in_month_column: bool = False,
) -> int:
    if marker and not first_title_in_month_column:
        marker_cell = sheet.cell(title_row, 1, marker)
        _style_ledger_cell(marker_cell, bold=True)
        marker_cell.alignment = Alignment(horizontal="center", vertical="center")

    for slot, item in enumerate(items):
        if item is None:
            continue
        title, _frame, _key_column, _key, _exempt, _blank = item
        value_start = SHARED_VALUE_START_COLUMNS[slot]
        if slot == 0 and first_title_in_month_column:
            title_start, title_end = 1, 4
            title_text = f"{marker}  {title}" if marker else title
        else:
            title_start = value_start
            title_end = value_start + 2
            title_text = title
        sheet.merge_cells(
            start_row=title_row,
            start_column=title_start,
            end_row=title_row,
            end_column=title_end,
        )
        title_cell = sheet.cell(title_row, title_start, title_text)
        _style_ledger_cell(title_cell, bold=True)
        title_cell.font = Font(name="맑은 고딕", size=10, bold=True, color=LEDGER_INK)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

    first_data_row = title_row + 1
    for month_index, month in enumerate(months):
        row = first_data_row + month_index
        month_cell = sheet.cell(row, 1, _month_label(month, months))
        _style_ledger_cell(month_cell)
        month_cell.alignment = Alignment(horizontal="center", vertical="bottom")
        for slot, item in enumerate(items):
            if item is None:
                continue
            _title, frame, key_column, key, exempt, blank = item
            fields: tuple[str | None, ...] = (
                "count",
                "supply_amount",
                None if exempt else "tax_amount",
            )
            for offset, field in enumerate(fields):
                cell = sheet.cell(row, SHARED_VALUE_START_COLUMNS[slot] + offset)
                if not blank and field is not None:
                    cell.value = _lookup(frame, key_column, key, month, field)
                _style_ledger_cell(cell)
                cell.alignment = Alignment(horizontal="right", vertical="bottom")
                cell.number_format = COUNT_FORMAT if field == "count" else MONEY_FORMAT

    total_row = first_data_row + len(months)
    total_label = sheet.cell(total_row, 1, "계")
    _style_ledger_cell(total_label, bold=True, total=True)
    total_label.alignment = Alignment(horizontal="center", vertical="bottom")
    for slot, item in enumerate(items):
        if item is None:
            continue
        _title, _frame, _key_column, _key, exempt, blank = item
        fields = ("count", "supply_amount", None if exempt else "tax_amount")
        for offset, field in enumerate(fields):
            column = SHARED_VALUE_START_COLUMNS[slot] + offset
            cell = sheet.cell(total_row, column)
            if not blank and field is not None:
                letter = get_column_letter(column)
                cell.value = f"=SUM({letter}{first_data_row}:{letter}{total_row - 1})"
            _style_ledger_cell(cell, bold=True, total=True)
            cell.alignment = Alignment(horizontal="right", vertical="bottom")
            cell.number_format = COUNT_FORMAT if field == "count" else MONEY_FORMAT
    return total_row


def _write_sales_type_summary(
    sheet,
    title_row: int,
    frame: pd.DataFrame,
) -> tuple[int, dict[str, str]]:
    label_column = 6
    value_start_column = 7
    components: list[tuple[str, str, str | None]] = []
    if any(_has_values(frame, "item", key, fields=("total_amount",)) for key in ("카드 과세", "카드 면세")):
        components.append(("카드", "카드 과세", "카드 면세"))
    if any(_has_values(frame, "item", key, fields=("total_amount",)) for key in ("현영 과세", "현영 면세")):
        components.append(("현영", "현영 과세", "현영 면세"))
    if _has_values(frame, "item", "제로페이", fields=("total_amount",)):
        components.append(("제로페이", "제로페이", None))

    if not components:
        return title_row - 1, {}

    has_taxable = any(_has_values(frame, "item", item, fields=("total_amount",)) for _, item, _ in components)
    has_exempt = any(
        exempt_item is not None
        and _has_values(frame, "item", exempt_item, fields=("total_amount",))
        for _, _, exempt_item in components
    )
    categories = [
        category
        for category, present in (("과세", has_taxable), ("면세", has_exempt))
        if present
    ]
    include_column_total = len(components) > 1
    total_column = value_start_column + len(components) if include_column_total else None

    header_label = sheet.cell(title_row, label_column)
    _style_ledger_cell(header_label, bold=True)
    header_label.alignment = Alignment(horizontal="center", vertical="center")
    for offset, (display, _, _) in enumerate(components):
        cell = sheet.cell(title_row, value_start_column + offset, display)
        _style_ledger_cell(cell, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if total_column is not None:
        cell = sheet.cell(title_row, total_column, "계")
        _style_ledger_cell(cell, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_refs: dict[str, str] = {}
    for row_offset, category in enumerate(categories, start=1):
        row_number = title_row + row_offset
        label = sheet.cell(row_number, label_column, category)
        _style_ledger_cell(label)
        label.alignment = Alignment(horizontal="center", vertical="center")
        for component_offset, (_, taxable_item, exempt_item) in enumerate(components):
            column = value_start_column + component_offset
            item = taxable_item if category == "과세" else exempt_item
            cell = sheet.cell(row_number, column)
            if item is not None and _has_values(frame, "item", item, fields=("total_amount",)):
                cell.value = _total_value(frame, "item", item, "total_amount")
            _style_ledger_cell(cell)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = MONEY_FORMAT
        if total_column is not None:
            first_letter = get_column_letter(value_start_column)
            last_letter = get_column_letter(value_start_column + len(components) - 1)
            total_cell = sheet.cell(
                row_number,
                total_column,
                f"=SUM({first_letter}{row_number}:{last_letter}{row_number})",
            )
            _style_ledger_cell(total_cell)
            total_cell.alignment = Alignment(horizontal="right", vertical="center")
            total_cell.number_format = MONEY_FORMAT
            row_refs[category] = f"{get_column_letter(total_column)}{row_number}"
        else:
            row_refs[category] = f"{get_column_letter(value_start_column)}{row_number}"

    last_row = title_row + len(categories)
    if len(categories) > 1:
        total_row = last_row + 1
        label = sheet.cell(total_row, label_column, "계")
        _style_ledger_cell(label, bold=True, total=True)
        label.alignment = Alignment(horizontal="center", vertical="center")
        last_value_column = total_column or value_start_column + len(components) - 1
        for column in range(value_start_column, last_value_column + 1):
            letter = get_column_letter(column)
            cell = sheet.cell(
                total_row,
                column,
                f"=SUM({letter}{title_row + 1}:{letter}{last_row})",
            )
            _style_ledger_cell(cell, bold=True, total=True)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = MONEY_FORMAT
        last_row = total_row
    return last_row, row_refs


def _build_summary(workbook: Workbook, result: ProcessingResult, settings: CompanySettings) -> None:
    sheet = workbook.active
    sheet.title = "집계표"
    months = sorted(result.transactions["month"].dropna().astype(str).unique())
    _set_summary_heading(sheet, settings.name, months)

    rows_per_band = len(months) + 3
    slots_per_band = len(SHARED_VALUE_START_COLUMNS)
    purchase_categories = [
        category
        for category in _ordered_categories(settings)
        if _has_values(result.purchase_by_category, "account_category", category)
    ]
    if _has_values(result.purchase_by_category, "account_category", "고정"):
        purchase_categories.insert(min(2, len(purchase_categories)), "고정")

    purchase_items: list[tuple[str, pd.DataFrame, str, str, bool, bool] | None] = [
        (category, result.purchase_by_category, "account_category", category, False, False)
        for category in purchase_categories
    ]
    invoice_present = _has_values(result.purchase_summary, "item", "세금계산서 매입계")
    if invoice_present:
        purchase_items.append(
            ("세금계산서 매입계", result.purchase_summary, "item", "세금계산서 매입계", False, False)
        )

    purchase_title_row = 3
    purchase_total_rows: list[int] = []
    for item_index in range(0, len(purchase_items), slots_per_band):
        title_row = purchase_title_row + (item_index // slots_per_band) * rows_per_band
        purchase_total_rows.append(
            _write_shared_month_band(
                sheet,
                title_row,
                purchase_items[item_index : item_index + slots_per_band],
                months,
                marker="①" if item_index == 0 else "",
            )
        )
    purchase_end_row = max(purchase_total_rows, default=1)

    if invoice_present:
        invoice_index = len(purchase_items) - 1
        invoice_slot = invoice_index % slots_per_band
        invoice_total_row = purchase_total_rows[invoice_index // slots_per_band]
        invoice_count_column = SHARED_VALUE_START_COLUMNS[invoice_slot]
        invoice_supply_column = invoice_count_column + 1
        invoice_tax_column = invoice_count_column + 2
        invoice_supply_ref = f"{get_column_letter(invoice_supply_column)}{invoice_total_row}"
        invoice_tax_ref = f"{get_column_letter(invoice_tax_column)}{invoice_total_row}"
        supply_letter = get_column_letter(invoice_supply_column)
        tax_letter = get_column_letter(invoice_tax_column)

        def write_summary_row(
            row_number: int,
            label: str,
            supply_value: object,
            tax_value: object,
            *,
            total: bool = False,
        ) -> None:
            for column, value in zip(
                (invoice_count_column, invoice_supply_column, invoice_tax_column),
                (label, supply_value, tax_value),
            ):
                cell = sheet.cell(row_number, column, value)
                _style_ledger_cell(cell, bold=True, total=total)
                cell.alignment = Alignment(horizontal="right", vertical="bottom")
                if column != invoice_count_column:
                    cell.number_format = MONEY_FORMAT

        adjustment_row = invoice_total_row + 1
        write_summary_row(
            adjustment_row,
            "단수차이 조정",
            f"={invoice_supply_ref}",
            f"=ROUNDDOWN({supply_letter}{adjustment_row}*0.1,0)",
        )
        base_supply_ref = f"{supply_letter}{adjustment_row}"
        base_tax_ref = f"{tax_letter}{adjustment_row}"
        summary_row = adjustment_row + 1

        if _has_values(result.purchase_summary, "item", "고정"):
            general_row = summary_row
            fixed_row = general_row + 1
            split_total_row = fixed_row + 1
            write_summary_row(
                general_row,
                "일반",
                _total_value(result.purchase_summary, "item", "일반매입", "supply_amount"),
                f"={tax_letter}{adjustment_row}-{tax_letter}{fixed_row}",
            )
            write_summary_row(
                fixed_row,
                "고정",
                _total_value(result.purchase_summary, "item", "고정", "supply_amount"),
                _total_value(result.purchase_summary, "item", "고정", "tax_amount"),
            )
            write_summary_row(
                split_total_row,
                "계",
                f"={supply_letter}{general_row}+{supply_letter}{fixed_row}",
                f"={tax_letter}{general_row}+{tax_letter}{fixed_row}",
                total=True,
            )
            base_supply_ref = f"{supply_letter}{split_total_row}"
            base_tax_ref = f"{tax_letter}{split_total_row}"
            summary_row = split_total_row + 1

        other_rows: list[int] = []
        for label, key in (("카드외", "카드외"), ("의제매입세액", "의제매입세액")):
            if _has_values(result.purchase_summary, "item", key):
                write_summary_row(
                    summary_row,
                    label,
                    _total_value(result.purchase_summary, "item", key, "supply_amount"),
                    _total_value(result.purchase_summary, "item", key, "tax_amount"),
                )
                other_rows.append(summary_row)
                summary_row += 1

        deduction_labels = [
            label
            for label in ("불공", "공통")
            if _has_values(result.purchase_summary, "item", label)
        ]
        overall_supply_ref = base_supply_ref
        overall_tax_ref = base_tax_ref
        if other_rows or deduction_labels:
            overall_row = summary_row
            write_summary_row(
                overall_row,
                "계",
                f"={base_supply_ref}" + "".join(
                    f"+{supply_letter}{row_number}" for row_number in other_rows
                ),
                f"={base_tax_ref}" + "".join(
                    f"+{tax_letter}{row_number}" for row_number in other_rows
                ),
                total=True,
            )
            overall_supply_ref = f"{supply_letter}{overall_row}"
            overall_tax_ref = f"{tax_letter}{overall_row}"
            summary_row += 1

        deduction_rows: list[int] = []
        for label in deduction_labels:
            write_summary_row(
                summary_row,
                label,
                _total_value(result.purchase_summary, "item", label, "supply_amount"),
                _total_value(result.purchase_summary, "item", label, "tax_amount"),
            )
            deduction_rows.append(summary_row)
            summary_row += 1
        if deduction_rows:
            write_summary_row(
                summary_row,
                "차감계",
                f"={overall_supply_ref}" + "".join(
                    f"-{supply_letter}{row_number}" for row_number in deduction_rows
                ),
                f"={overall_tax_ref}" + "".join(
                    f"-{tax_letter}{row_number}" for row_number in deduction_rows
                ),
                total=True,
            )
            summary_row += 1
        purchase_end_row = max(purchase_end_row, summary_row - 1)

    detail_items: list[tuple[str, pd.DataFrame, str, str, bool, bool] | None] = []
    for label in ("카과", "현과"):
        if _has_values(result.purchase_summary, "item", label):
            detail_items.append((label, result.purchase_summary, "item", label, False, False))
    if _has_values(result.purchase_summary, "item", "의제매입세액"):
        detail_items.append(
            ("의제매입세액", result.purchase_summary, "item", "의제매입세액", False, False)
        )
    for label in ("불공", "공통"):
        if _has_values(result.purchase_summary, "item", label):
            detail_items.append((label, result.purchase_summary, "item", label, False, False))
    if _has_values(result.purchase_summary, "item", "면세 매입"):
        detail_items.append(
            ("면세 매입", result.purchase_summary, "item", "면세 매입", True, False)
        )

    detail_end_row = purchase_end_row
    if detail_items:
        detail_title_row = purchase_end_row + 4
        for item_index in range(0, len(detail_items), slots_per_band):
            title_row = detail_title_row + (item_index // slots_per_band) * rows_per_band
            detail_end_row = _write_shared_month_band(
                sheet,
                title_row,
                detail_items[item_index : item_index + slots_per_band],
                months,
                marker="②" if item_index == 0 else "",
            )

    sales_accounts = [
        account
        for account in sorted(result.sales_by_account["account_name"].dropna().astype(str).unique())
        if _has_values(result.sales_by_account, "account_name", account)
    ]
    sales_items: list[tuple[str, pd.DataFrame, str, str, bool, bool] | None] = [
        (
            f"{account} · 세금계산서 매출",
            result.sales_by_account,
            "account_name",
            account,
            False,
            False,
        )
        for account in sales_accounts
    ]
    if _has_values(result.sales_summary, "item", "면세 계산서 매출"):
        sales_items.append(
            (
                "면세 계산서 매출",
                result.sales_summary,
                "item",
                "면세 계산서 매출",
                True,
                False,
            )
        )

    sales_title_row = detail_end_row + 6
    sales_end_row = sales_title_row - 1
    taxable_sales_refs: list[str] = []
    for item_index in range(0, len(sales_items), slots_per_band):
        title_row = sales_title_row + (item_index // slots_per_band) * rows_per_band
        band_items = sales_items[item_index : item_index + slots_per_band]
        band_total_row = _write_shared_month_band(
            sheet,
            title_row,
            band_items,
            months,
            marker="③" if item_index == 0 else "",
            first_title_in_month_column=item_index == 0,
        )
        sales_end_row = max(sales_end_row, band_total_row)
        for slot, item in enumerate(band_items):
            if item is not None and item[2] == "account_name":
                supply_column = SHARED_VALUE_START_COLUMNS[slot] + 1
                taxable_sales_refs.append(
                    f"{get_column_letter(supply_column)}{band_total_row}"
                )

    row = max(purchase_end_row, detail_end_row, sales_end_row)
    sales_adjustment_row: int | None = None
    if taxable_sales_refs:
        sales_adjustment_row = sales_end_row + 1
        label = sheet.cell(sales_adjustment_row, 2, "단수차이 조정")
        supply = sheet.cell(sales_adjustment_row, 3, "=" + "+".join(taxable_sales_refs))
        tax = sheet.cell(
            sales_adjustment_row,
            4,
            f"=ROUNDDOWN(C{sales_adjustment_row}*0.1,0)",
        )
        for cell in (label, supply, tax):
            _style_ledger_cell(cell, bold=True)
            cell.alignment = Alignment(horizontal="right", vertical="bottom")
            cell.number_format = MONEY_FORMAT
        row = max(row, sales_adjustment_row)

    card_items = [
        (display, key)
        for display, key in (
            ("카드 합계\n(카과+카면+카영)", "카드매출"),
            ("현영 합계\n(현과+현면+현영)", "현영매출"),
            ("제로페이", "제로페이"),
        )
        if _has_values(result.sales_summary, "item", key, fields=("total_amount",))
    ]
    if card_items:
        card_outside_row = (sales_adjustment_row or sales_end_row) + 1
        card_table_title_row = card_outside_row + 5
        card_total_row, _ = _write_total_amount_table(
            sheet,
            card_table_title_row,
            card_items,
            result.sales_summary,
            months,
        )
        sales_type_end_row, sales_type_refs = _write_sales_type_summary(
            sheet,
            card_table_title_row,
            result.sales_summary,
        )

        label = sheet.cell(card_outside_row, 2, "카드외")
        _style_ledger_cell(label, bold=True)
        label.alignment = Alignment(horizontal="right", vertical="bottom")
        taxable_ref = sales_type_refs.get("과세")
        if taxable_ref:
            supply = sheet.cell(card_outside_row, 3, f"=ROUND({taxable_ref}/1.1,0)")
            tax = sheet.cell(card_outside_row, 4, f"={taxable_ref}-C{card_outside_row}")
            for cell in (supply, tax):
                _style_ledger_cell(cell, bold=True)
                cell.alignment = Alignment(horizontal="right", vertical="bottom")
                cell.number_format = MONEY_FORMAT
        exempt_ref = sales_type_refs.get("면세")
        if exempt_ref:
            exempt = sheet.cell(card_outside_row, 6, f"={exempt_ref}")
            _style_ledger_cell(exempt, bold=True)
            exempt.alignment = Alignment(horizontal="right", vertical="bottom")
            exempt.number_format = MONEY_FORMAT
        row = max(row, card_total_row, sales_type_end_row)

    for row_number in range(1, row + 1):
        sheet.row_dimensions[row_number].height = LEDGER_ROW_HEIGHT

    for row_cells in sheet.iter_rows(
        min_row=1,
        max_row=row,
        min_col=1,
        max_col=SUMMARY_END_COLUMN,
    ):
        for cell in row_cells:
            cell.fill = PatternFill("solid", fgColor=WHITE)

    for row_number in range(3, row + 1):
        has_content = any(
            sheet.cell(row_number, column).value is not None
            for column in range(1, SUMMARY_END_COLUMN + 1)
        )
        for column in range(1, SUMMARY_END_COLUMN + 1):
            cell = sheet.cell(row_number, column)
            is_total_cell = getattr(cell.border.top, "style", None) == "medium"
            sheet.cell(row_number, column).border = Border(
                top=LEDGER_TOTAL_RULE if is_total_cell else None,
                bottom=LEDGER_RULE if has_content else None,
            )

    sheet.sheet_view.showGridLines = False
    widths = (4, 16, 14, 10, 11, 14, 10, 9, 14, 10, 14, 14, 10)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.paperWidth = "170mm"
    sheet.page_setup.paperHeight = "240mm"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.15
    sheet.page_margins.right = 0.15
    sheet.page_margins.top = 0.2
    sheet.page_margins.bottom = 0.2
    sheet.page_margins.header = 0
    sheet.page_margins.footer = 0
    sheet.print_options.horizontalCentered = True
    sheet.print_area = f"A1:{get_column_letter(SUMMARY_END_COLUMN)}{row}"


def _write_table_sheet(
    workbook: Workbook,
    name: str,
    title: str,
    headers: list[tuple[str, str]],
    frame: pd.DataFrame,
    *,
    alert_column: str | None = None,
) -> None:
    sheet = workbook.create_sheet(name)
    _set_title(sheet, title, len(headers))
    for column, (_, label) in enumerate(headers, start=1):
        cell = sheet.cell(3, column, label)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="맑은 고딕", bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_number, (_, record) in enumerate(frame.iterrows(), start=4):
        for column, (field, _) in enumerate(headers, start=1):
            value = _excel_value(record.get(field, ""))
            cell = sheet.cell(row_number, column, value)
            cell.border = Border(bottom=THIN_GRAY)
            cell.alignment = Alignment(vertical="top", wrap_text=field in {"item", "candidate_reason", "nondeductible_reason", "review_memo", "detail"})
            if field in {"supply_amount", "tax_amount", "total_amount", "expected", "actual", "difference"}:
                cell.number_format = "#,##0;[Red]-#,##0"
            if field == "date" and isinstance(value, (date, datetime)):
                cell.number_format = "yyyy-mm-dd"
        if alert_column and str(record.get(alert_column, "")) in {"실패", "판단 보류", "미분류"}:
            for cell in sheet[row_number]:
                cell.fill = PatternFill("solid", fgColor=RED)

    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{max(3, 3 + len(frame))}"
    sheet.sheet_view.showGridLines = False
    for index, (field, label) in enumerate(headers, start=1):
        width = max(12, min(42, len(label) * 2 + 4))
        if field in {"vendor", "item", "candidate_reason", "nondeductible_reason", "review_memo", "detail"}:
            width = 28 if field in {"vendor", "item"} else 36
        sheet.column_dimensions[get_column_letter(index)].width = width


def export_workbook(
    result: ProcessingResult,
    settings: CompanySettings,
    output_path: str | Path,
) -> Path:
    """처리 결과를 정본에 정의된 다섯 개 시트로 저장합니다."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    _build_summary(workbook, result, settings)

    review_headers = [
        ("date", "날짜"),
        ("vendor", "거래처"),
        ("item", "품명"),
        ("account_code", "계정코드"),
        ("account_name", "계정과목"),
        ("supply_amount", "공급가액"),
        ("tax_amount", "세액"),
        ("original_type", "원본 유형"),
        ("candidate_reason", "후보 근거"),
        ("review_status", "최종 판정"),
        ("final_type", "최종 유형"),
        ("nondeductible_reason", "불공 사유"),
        ("review_memo", "메모"),
    ]
    _write_table_sheet(
        workbook,
        "불공검토",
        "불공 검토",
        review_headers,
        result.review,
        alert_column="review_status",
    )

    detail_headers = [
        ("row_id", "원본 행 ID"),
        ("date", "전표일자"),
        ("division", "매입·매출"),
        ("month", "월"),
        ("vendor", "거래처"),
        ("item", "품명"),
        ("supply_amount", "공급가액"),
        ("tax_amount", "세액"),
        ("total_amount", "합계금액"),
        ("account_code", "계정코드"),
        ("account_name", "계정과목"),
        ("account_category", "계정 분류"),
        ("original_type", "원본 유형"),
        ("final_type", "최종 유형"),
        ("nondeductible", "불공 여부"),
        ("nondeductible_reason", "불공 사유"),
        ("candidate_reason", "후보 근거"),
        ("review_status", "최종 판정"),
        ("review_memo", "메모"),
    ]
    _write_table_sheet(workbook, "거래상세", "원본 상세 거래 및 파생 값", detail_headers, result.transactions)

    unclassified = result.transactions[result.transactions["account_category"].eq("미분류")]
    _write_table_sheet(
        workbook,
        "미분류",
        "계정 미분류 거래",
        detail_headers,
        unclassified,
        alert_column="account_category",
    )

    validation_headers = [
        ("check", "검산 항목"),
        ("status", "상태"),
        ("expected", "기대값"),
        ("actual", "결과값"),
        ("difference", "차이"),
        ("detail", "확인 사항"),
    ]
    _write_table_sheet(
        workbook,
        "검산",
        f"검산 결과 · {'완료' if result.validation_passed else '실패'}",
        validation_headers,
        result.validation,
        alert_column="status",
    )
    workbook.save(path)
    return path
