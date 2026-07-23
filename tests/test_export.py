# 결과 엑셀의 필수 시트와 핵심 집계 셀 및 수식을 검증합니다.
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook
import pandas as pd

from tally import CompanySettings, export_workbook, process_transactions


def test_export_creates_required_sheets_and_summary_formulas(tmp_path) -> None:
    rows = [
        {
            "row_id": "Sheet1:2",
            "sheet": "Sheet1",
            "source_row": 2,
            "division": "매입",
            "date": date(2026, 4, 1),
            "month": "2026-04",
            "vendor": "상사",
            "item": "재료",
            "supply_amount": Decimal(1000),
            "tax_amount": Decimal(100),
            "total_amount": Decimal(1100),
            "original_type": "과세",
            "account_code": "146",
            "account_name": "상품",
            "card_company": "",
            "card_number": "",
        },
        {
            "row_id": "Sheet1:3",
            "sheet": "Sheet1",
            "source_row": 3,
            "division": "매입",
            "date": date(2026, 4, 2),
            "month": "2026-04",
            "vendor": "상사",
            "item": "접대",
            "supply_amount": Decimal(500),
            "tax_amount": Decimal(50),
            "total_amount": Decimal(550),
            "original_type": "불공",
            "account_code": "813",
            "account_name": "접대비",
            "card_company": "",
            "card_number": "",
        },
        {
            "row_id": "Sheet1:4",
            "sheet": "Sheet1",
            "source_row": 4,
            "division": "매출",
            "date": date(2026, 4, 3),
            "month": "2026-04",
            "vendor": "고객",
            "item": "매출",
            "supply_amount": Decimal(2000),
            "tax_amount": Decimal(200),
            "total_amount": Decimal(2200),
            "original_type": "카과",
            "account_code": "401",
            "account_name": "상품매출",
            "card_company": "국민",
            "card_number": "1234",
        },
    ]
    settings = CompanySettings(name="테스트상사")
    result = process_transactions(pd.DataFrame(rows), settings)
    output = export_workbook(result, settings, tmp_path / "result.xlsx")

    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == ["집계표", "불공검토", "거래상세", "미분류", "검산"]
    summary = workbook["집계표"]
    assert summary["A1"].value == "2026년 1기 확정"
    assert summary["E1"].value == "테스트상사"
    assert summary["I1"].value == "부가세 집계표"
    assert summary["A3"].value == "①"
    assert summary["B3"].value == "상품"
    assert summary["E3"].value == "기타"
    assert summary["H3"].value == "세금계산서 매입계"
    assert summary["A4"].value == "4월"
    assert summary["B4"].value == 1
    assert summary["C4"].value == 1000
    assert summary["D4"].value == 100
    assert summary["E4"].value == 1
    assert summary["F4"].value == 500
    assert summary["G4"].value == 50
    assert summary["H4"].value == 2
    assert summary["I4"].value == 1500
    assert summary["J4"].value == 150
    assert summary["B4"].number_format == '"("0")"'
    assert summary.row_dimensions[2].height == 25
    assert summary.row_dimensions[3].height == 25
    assert summary.row_dimensions[4].height == 25
    assert summary["B4"].alignment.vertical == "bottom"
    assert summary["B4"].border.bottom.style == "dotted"
    assert summary["A5"].value == "계"
    assert summary["B5"].value == "=SUM(B4:B4)"
    assert summary.row_dimensions[5].height == 25
    assert summary["J5"].border.bottom.style == "dotted"
    assert summary["J5"].border.top.style == "medium"
    assert summary["H6"].value == "단수차이 조정"
    assert summary["I6"].value == "=I5"
    assert summary["J6"].value == "=ROUNDDOWN(I6*0.1,0)"
    assert summary.row_dimensions[6].height == 25
    assert summary["I6"].alignment.vertical == "bottom"
    assert summary["H7"].value == "불공"
    assert summary["I7"].value == 500
    assert summary["J7"].value == 50
    assert summary["H8"].value == "차감계"
    assert summary["I8"].value == "=I6-I7"
    assert summary["F9"].value == "납부"
    assert summary["G9"].value == "=D20-J8"
    assert summary["B12"].value == "불공"
    assert summary["B13"].value == 1
    assert summary["B20"].value == "카드외"
    assert summary["C20"].value == "=ROUNDUP(G24/1.1,0)"
    assert summary["D20"].value == "=G24-C20"
    assert summary["B23"].value == "카드"
    assert summary["B24"].value == 2200
    assert summary["B25"].value == "=SUM(B24:B24)"
    assert summary["G23"].value == "카드"
    assert summary["F24"].value == "과세"
    assert summary["G24"].value == 2200
    assert not any(
        cell.value
        in {"원재료(도급)", "제조경비", "도급경비", "고정", "현과", "면세"}
        for row in summary.iter_rows()
        for cell in row
    )
    assert summary["A3"].fill.fill_type == "solid"
    assert summary["A3"].fill.fgColor.rgb.endswith("FFFFFF")
    assert summary["A3"].border.bottom.style == "dotted"
    assert summary.print_area == "'집계표'!$A$1:$J$25"
    assert summary.page_setup.orientation == "portrait"
    assert summary.page_setup.paperWidth == "170mm"
    assert summary.page_setup.paperHeight == "240mm"
    assert summary.page_setup.fitToWidth == 1
    assert summary.page_setup.fitToHeight == 1
    assert any(
        isinstance(cell.value, str) and cell.value.startswith("=")
        for row in summary.iter_rows()
        for cell in row
    )
    assert workbook["거래상세"].max_row == 6
    assert workbook["거래상세"].max_column == 23
    assert workbook["거래상세"]["B3"].value == "원본 시트"
    assert workbook["거래상세"]["C3"].value == "원본 행"
    assert workbook["거래상세"]["I3"].value == "카드사"
    assert workbook["거래상세"]["J3"].value == "카드번호"
    assert workbook["거래상세"]["B4"].value == "Sheet1"
    assert workbook["거래상세"]["C4"].value == 2
    assert workbook["검산"]["A1"].value.startswith("검산 결과")


def test_export_places_other_input_tax_and_deductions_before_sales(tmp_path) -> None:
    def row(
        number: int,
        division: str,
        original_type: str,
        supply: int,
        tax: int,
        code: str = "",
        account_name: str = "",
    ) -> dict[str, object]:
        return {
            "row_id": f"Sheet1:{number}",
            "sheet": "Sheet1",
            "source_row": number,
            "division": division,
            "date": date(2026, 4, number),
            "month": "2026-04",
            "vendor": "거래처",
            "item": "품목",
            "supply_amount": Decimal(supply),
            "tax_amount": Decimal(tax),
            "total_amount": Decimal(supply + tax),
            "original_type": original_type,
            "account_code": code,
            "account_name": account_name,
            "card_company": "",
            "card_number": "",
        }

    rows = [
        row(1, "매입", "과세", 1000, 100, "146", "상품"),
        row(2, "매입", "불공", 200, 20, "813", "접대비"),
        row(3, "매입", "공통", 150, 15, "813", "공통매입"),
        row(4, "매입", "카과", 300, 30),
        row(5, "매입", "현과", 400, 40),
        row(6, "매입", "의제매입세액", 50, 5),
        row(7, "매출", "과세", 1000, 100, "401", "상품매출"),
    ]
    settings = CompanySettings(name="테스트상사")
    result = process_transactions(pd.DataFrame(rows), settings)
    output = export_workbook(result, settings, tmp_path / "vat-layout.xlsx")

    summary = load_workbook(output, data_only=False)["집계표"]
    assert summary["H3"].value == "세금계산서 매입계"
    assert summary["H6"].value == "단수차이 조정"
    assert summary["H7"].value == "카드외"
    assert summary["I7"].value == 700
    assert summary["H8"].value == "의제매입세액"
    assert summary["I8"].value == 50
    assert summary["H9"].value == "계"
    assert summary["I9"].value == "=I6+I7+I8"
    assert summary["H10"].value == "불공"
    assert summary["I10"].value == 200
    assert summary["H11"].value == "공통"
    assert summary["I11"].value == 150
    assert summary["H12"].value == "차감계"
    assert summary["I12"].value == "=I9-I10-I11"
    assert summary["A16"].value == "②"
    assert summary["B16"].value == "카과"
    assert summary["E16"].value == "현과"
    assert summary["H16"].value == "의제매입세액"
    assert summary["B17"].value == 1
    assert summary["C17"].value == 300
    assert summary["D17"].value == 30
    assert summary["E17"].value == 1
    assert summary["F17"].value == 400
    assert summary["G17"].value == 40
    assert summary["A24"].value == "③  상품매출 · 세금계산서 매출"
    assert summary["E24"].value == "불공"
    assert summary["H24"].value == "공통"


def test_purchase_rounding_with_only_deduction_skips_intermediate_total(tmp_path) -> None:
    rows = []
    for number, original_type, code, supply, tax in (
        (1, "과세", "146", 800, 79),
        (2, "불공", "813", 200, 20),
    ):
        rows.append(
            {
                "row_id": f"Sheet1:{number}",
                "sheet": "Sheet1",
                "source_row": number,
                "division": "매입",
                "date": date(2026, 4, number),
                "month": "2026-04",
                "vendor": "거래처",
                "item": "품목",
                "supply_amount": Decimal(supply),
                "tax_amount": Decimal(tax),
                "total_amount": Decimal(supply + tax),
                "original_type": original_type,
                "account_code": code,
                "account_name": "상품" if code == "146" else "접대비",
                "card_company": "",
                "card_number": "",
            }
        )

    settings = CompanySettings(name="중간계제외")
    result = process_transactions(pd.DataFrame(rows), settings)
    output = export_workbook(result, settings, tmp_path / "no-duplicate-total.xlsx")
    summary = load_workbook(output, data_only=False)["집계표"]

    assert summary["H6"].value == "단수차이 조정"
    assert summary["H7"].value == "불공"
    assert summary["H8"].value == "차감계"
    assert summary["I8"].value == "=I6-I7"


def test_export_splits_fixed_assets_after_rounding_adjustment(tmp_path) -> None:
    rows = [
        {
            "row_id": "Sheet1:2",
            "sheet": "Sheet1",
            "source_row": 2,
            "division": "매입",
            "date": date(2026, 4, 1),
            "month": "2026-04",
            "vendor": "일반거래처",
            "item": "일반",
            "supply_amount": Decimal(1005),
            "tax_amount": Decimal(100),
            "total_amount": Decimal(1105),
            "original_type": "과세",
            "account_code": "146",
            "account_name": "상품",
            "card_company": "",
            "card_number": "",
        },
        {
            "row_id": "Sheet1:3",
            "sheet": "Sheet1",
            "source_row": 3,
            "division": "매입",
            "date": date(2026, 4, 2),
            "month": "2026-04",
            "vendor": "고정거래처",
            "item": "비품",
            "supply_amount": Decimal(1005),
            "tax_amount": Decimal(100),
            "total_amount": Decimal(1105),
            "original_type": "과세",
            "account_code": "212",
            "account_name": "비품",
            "card_company": "",
            "card_number": "",
        },
    ]
    settings = CompanySettings(name="고정테스트")
    result = process_transactions(pd.DataFrame(rows), settings)
    output = export_workbook(result, settings, tmp_path / "fixed-layout.xlsx")

    summary = load_workbook(output, data_only=False)["집계표"]
    assert summary["E3"].value == "고정"
    assert summary["H3"].value == "세금계산서 매입계"
    assert summary["H6"].value == "단수차이 조정"
    assert summary["I6"].value == "=I5"
    assert summary["J6"].value == "=ROUNDDOWN(I6*0.1,0)"
    assert summary["H7"].value == "일반"
    assert summary["I7"].value == 1005
    assert summary["J7"].value == "=J6-J8"
    assert summary["H8"].value == "고정"
    assert summary["I8"].value == 1005
    assert summary["J8"].value == 100
    assert summary["H9"].value == "계"
    assert summary["I9"].value == "=I7+I8"
    assert summary["J9"].value == "=J7+J8"
    assert summary["H10"].value is None


def test_export_matches_tobacco_and_card_cash_sales_layout(tmp_path) -> None:
    def row(
        number: int,
        division: str,
        original_type: str,
        supply: int,
        tax: int,
        *,
        vendor: str = "거래처",
        code: str = "401",
        account_name: str = "상품매출",
    ) -> dict[str, object]:
        return {
            "row_id": f"Sheet1:{number}",
            "sheet": "Sheet1",
            "source_row": number,
            "division": division,
            "date": date(2026, 4, number),
            "month": "2026-04",
            "vendor": vendor,
            "item": "품목",
            "supply_amount": Decimal(supply),
            "tax_amount": Decimal(tax),
            "total_amount": Decimal(supply + tax),
            "original_type": original_type,
            "account_code": code,
            "account_name": account_name,
            "card_company": "제로페이" if "제로페이" in vendor else "",
            "card_number": "",
        }

    rows = [
        row(1, "매입", "과세", 1000, 100, vendor="일반상사", code="146", account_name="상품"),
        row(2, "매입", "과세", 2000, 200, vendor="케이티앤지", code="146", account_name="상품"),
        row(3, "매입", "과세", 3000, 300, vendor="기타상사", code="813", account_name="기타"),
        row(4, "매출", "과세", 1000, 100),
        row(5, "매출", "면세", 500, 0),
        row(6, "매출", "카과", 1000, 100),
        row(7, "매출", "카과", 200, 20, vendor="제로페이"),
        row(8, "매출", "카면", 300, 0),
        row(9, "매출", "카영", 400, 0),
        row(10, "매출", "현과", 500, 50),
        row(11, "매출", "현면", 200, 0),
        row(12, "매출", "현영", 100, 0),
    ]
    result = process_transactions(pd.DataFrame(rows), CompanySettings(name="영진상회"))
    output = export_workbook(result, CompanySettings(name="영진상회"), tmp_path / "youngjin-layout.xlsx")

    summary = load_workbook(output, data_only=False)["집계표"]
    assert summary["B3"].value == "상품"
    assert summary["E3"].value == "담배"
    assert summary["H3"].value == "기타"
    assert summary["K3"].value == "세금계산서 매입계"
    assert summary["B4"].value == 1
    assert summary["E4"].value == 1
    assert summary["F4"].value == 2000
    assert summary["K6"].value == "단수차이 조정"
    assert summary["L6"].value == "=L5"
    assert summary["M6"].value == "=ROUNDDOWN(L6*0.1,0)"

    assert summary["A12"].value == "③  상품매출 · 세금계산서 매출"
    assert summary["B13"].value == 1
    assert summary["C13"].value == 1000
    assert summary["E12"].value == "면세 계산서 매출"
    assert summary["E13"].value == 1
    assert summary["F13"].value == 500
    assert summary["B15"].value == "단수차이 조정"
    assert summary["C15"].value == "=C14"
    assert summary["D15"].value == "=ROUNDDOWN(C15*0.1,0)"
    assert summary["B16"].value == "카드외"
    assert summary["C16"].value == "=ROUNDUP(J21/1.1,0)"
    assert summary["D16"].value == "=J21-C16"
    assert summary["F16"].value == "=J22"
    assert summary["B17"].value == "계"

    assert summary["B20"].value == "카드"
    assert summary["C20"].value == "현영"
    assert summary["D20"].value == "제로페이"
    assert summary["B21"].value == 1800
    assert summary["C21"].value == 850
    assert summary["D21"].value == 220
    assert summary["B22"].value == "=SUM(B21:B21)"

    assert summary["G20"].value == "카드"
    assert summary["H20"].value == "현영"
    assert summary["I20"].value == "제로페이"
    assert summary["J20"].value == "계"
    assert summary["F21"].value == "과세"
    assert summary["G21"].value == 1100
    assert summary["H21"].value == 550
    assert summary["I21"].value == 220
    assert summary["J21"].value == "=SUM(G21:I21)"
    assert summary["F22"].value == "면세"
    assert summary["G22"].value == 700
    assert summary["H22"].value == 300
    assert summary["I22"].value is None
    assert summary["J22"].value == "=SUM(G22:I22)"
    assert summary["F23"].value == "계"
    assert summary["G23"].value == "=SUM(G21:G22)"
    assert summary["J23"].value == "=SUM(J21:J22)"
    assert summary["G20"].font.name == "맑은 고딕"
    assert summary["G20"].font.sz == 9
    assert summary["G20"].font.bold
    assert summary["G21"].font.name == "맑은 고딕"
    assert summary["G21"].font.sz == 9
    assert summary["F23"].border.top.style == "medium"
    assert summary.print_area == "'집계표'!$A$1:$M$23"


def test_export_uses_sample_layout_for_six_month_custom_category(tmp_path) -> None:
    rows = []
    row_number = 1
    for month in range(1, 7):
        for original_type, code, account_name, supply, tax in (
            ("과세", "899", "운송경비", 100, 10),
            ("과세", "813", "기타", 50, 5),
            ("과세", "212", "차량운반구", 1000, 100),
            ("카과", "", "", 200, 20),
            ("현과", "", "", 100, 10),
            ("과세", "401", "차량운반구", 5000, 500),
        ):
            division = "매출" if code == "401" else "매입"
            rows.append(
                {
                    "row_id": f"Sheet1:{row_number}",
                    "sheet": "Sheet1",
                    "source_row": row_number,
                    "division": division,
                    "date": date(2026, month, 1),
                    "month": f"2026-{month:02d}",
                    "vendor": "거래처",
                    "item": "품목",
                    "supply_amount": Decimal(supply),
                    "tax_amount": Decimal(tax),
                    "total_amount": Decimal(supply + tax),
                    "original_type": original_type,
                    "account_code": code,
                    "account_name": account_name,
                    "card_company": "",
                    "card_number": "",
                }
            )
            row_number += 1

    settings = CompanySettings(
        name="화천택시",
        fixed_asset_codes={"212"},
        account_overrides={"899": "운송경비"},
    )
    result = process_transactions(pd.DataFrame(rows), settings)
    output = export_workbook(result, settings, tmp_path / "six-month-layout.xlsx")
    summary = load_workbook(output, data_only=False)["집계표"]

    assert summary["B3"].value == "운송경비"
    assert summary["E3"].value == "기타"
    assert summary["H3"].value == "고정"
    assert summary["K3"].value == "세금계산서 매입계"
    assert summary["K11"].value == "단수차이 조정"
    assert summary["A13"].value == "②"
    assert summary["B13"].value == "카과"
    assert summary["E13"].value == "현과"
    assert summary["A24"].value == "③  차량운반구 · 세금계산서 매출"
    assert summary.print_area == "'집계표'!$A$1:$M$32"


def test_export_includes_zero_rated_other_sales_and_optional_adjustments(tmp_path) -> None:
    def row(
        number: int,
        division: str,
        original_type: str,
        supply: int,
        tax: int,
        account_name: str,
    ) -> dict[str, object]:
        return {
            "row_id": f"Sheet1:{number}",
            "sheet": "Sheet1",
            "source_row": number,
            "division": division,
            "date": date(2026, 4, number),
            "month": "2026-04",
            "vendor": "거래처",
            "item": "품목",
            "supply_amount": Decimal(supply),
            "tax_amount": Decimal(tax),
            "total_amount": Decimal(supply + tax),
            "original_type": original_type,
            "account_code": "146" if division == "매입" else "401",
            "account_name": account_name,
            "card_company": "",
            "card_number": "",
        }

    rows = [
        row(1, "매입", "과세", 1000, 99, "상품"),
        row(2, "매입", "영세", 200, 0, "상품"),
        row(3, "매출", "과세", 1000, 99, "회사설정계정과목"),
        row(4, "매출", "건별", 500, 50, "상품매출"),
        row(5, "매출", "면건", 300, 0, "상품매출"),
    ]
    settings = CompanySettings(
        name="조정테스트",
        prior_period_credit=100,
        card_sales_deduction=20,
    )
    result = process_transactions(pd.DataFrame(rows), settings)
    output = export_workbook(result, settings, tmp_path / "adjusted-layout.xlsx")
    summary = load_workbook(output, data_only=False)["집계표"]

    labels = {
        cell.value: cell
        for row_cells in summary.iter_rows()
        for cell in row_cells
        if isinstance(cell.value, str) and not cell.value.startswith("=")
    }
    assert any(label.endswith("수입금액 · 세금계산서 매출") for label in labels)
    assert labels["영세 매입"]
    assert labels["기타"]
    other_row = labels["기타"].row
    assert summary.cell(other_row, 3).value == 500
    assert summary.cell(other_row, 4).value == 50
    assert summary.cell(other_row, 6).value == 300

    payable = labels["납부"]
    assert summary.cell(payable.row, payable.column + 1).value.startswith("=D")
    assert summary.cell(labels["카드"].row, labels["카드"].column + 1).value == 20
    assert summary.cell(labels["예정미환급"].row, labels["예정미환급"].column + 1).value == 100
    assert summary.cell(labels["차감"].row, labels["차감"].column + 1).value.startswith("=")
    assert sum(
        cell.value == "단수차이 조정"
        for row_cells in summary.iter_rows()
        for cell in row_cells
    ) == 2
    assert summary.print_area.endswith(f"$J${summary.max_row}")


def test_export_handles_all_payment_purchase_types_and_multiple_invoice_accounts(
    tmp_path,
) -> None:
    def row(
        number: int,
        division: str,
        original_type: str,
        supply: int,
        tax: int,
        *,
        code: str = "",
        account_name: str = "",
    ) -> dict[str, object]:
        return {
            "row_id": f"Sheet1:{number}",
            "sheet": "Sheet1",
            "source_row": number,
            "division": division,
            "date": date(2026, 4, number),
            "month": "2026-04",
            "vendor": "거래처",
            "item": "품목",
            "supply_amount": Decimal(supply),
            "tax_amount": Decimal(tax),
            "total_amount": Decimal(supply + tax),
            "original_type": original_type,
            "account_code": code,
            "account_name": account_name,
            "card_company": "",
            "card_number": "",
        }

    rows = [
        row(1, "매입", "과세", 1000, 100, code="146", account_name="상품"),
        row(2, "매입", "카면", 500, 0),
        row(3, "매입", "카영", 100, 0),
        row(4, "매입", "현면", 200, 0),
        row(5, "매입", "현영", 50, 0),
        row(6, "매출", "과세", 2000, 200, code="401", account_name="상품매출"),
        row(7, "매출", "과세", 3000, 300, code="206", account_name="차량운반구"),
    ]
    settings = CompanySettings(name="복합유형")
    result = process_transactions(pd.DataFrame(rows), settings)
    output = export_workbook(result, settings, tmp_path / "generic-layout.xlsx")
    summary = load_workbook(output, data_only=False)["집계표"]

    values = [
        cell.value
        for row_cells in summary.iter_rows()
        for cell in row_cells
        if cell.value is not None
    ]
    for label in ("카면", "카영", "현면", "현영"):
        assert label in values
    assert values.count("③  세금계산서 매출") == 1
    assert "상품매출 · 세금계산서 매출" not in values
    assert "차량운반구 · 세금계산서 매출" not in values

    title = next(
        cell
        for row_cells in summary.iter_rows()
        for cell in row_cells
        if cell.value == "③  세금계산서 매출"
    )
    assert summary.cell(title.row, 5).value == "상품매출"
    assert summary.cell(title.row, 6).value == "차량운반구"
    assert summary.cell(title.row + 1, 3).value == 5000
    assert summary.cell(title.row + 1, 5).value == 2000
    assert summary.cell(title.row + 1, 6).value == 3000
    assert summary.cell(title.row + 3, 3).value == f"=C{title.row + 2}"


def test_export_keeps_payment_outside_three_wide_purchase_detail_tables(
    tmp_path,
) -> None:
    def row(
        number: int,
        division: str,
        original_type: str,
        supply: int,
        tax: int,
        *,
        vendor: str = "거래처",
        code: str = "",
        account_name: str = "",
    ) -> dict[str, object]:
        return {
            "row_id": f"Sheet1:{number}",
            "sheet": "Sheet1",
            "source_row": number,
            "division": division,
            "date": date(2026, 4, number),
            "month": "2026-04",
            "vendor": vendor,
            "item": "품목",
            "supply_amount": Decimal(supply),
            "tax_amount": Decimal(tax),
            "total_amount": Decimal(supply + tax),
            "original_type": original_type,
            "account_code": code,
            "account_name": account_name,
            "card_company": "",
            "card_number": "",
        }

    rows = [
        row(1, "매입", "과세", 1000, 100, code="146", account_name="상품"),
        row(
            2,
            "매입",
            "과세",
            2000,
            200,
            vendor="케이티앤지",
            code="146",
            account_name="상품",
        ),
        row(3, "매입", "과세", 3000, 300, code="813", account_name="기타"),
        row(4, "매입", "카면", 400, 0),
        row(5, "매입", "면세", 500, 0),
        row(6, "매입", "영세", 600, 0),
        row(7, "매출", "과세", 7000, 700, code="401", account_name="상품매출"),
    ]
    settings = CompanySettings(name="넓은배치")
    result = process_transactions(pd.DataFrame(rows), settings)
    output = export_workbook(result, settings, tmp_path / "wide-detail-payment.xlsx")
    summary = load_workbook(output, data_only=False)["집계표"]

    assert summary["A7"].value == "②"
    assert summary["B7"].value == "카면"
    assert summary["E7"].value == "면세 매입"
    assert summary["H7"].value == "영세 매입"
    assert summary["L7"].value == "납부"
    assert str(summary["M7"].value).startswith("=")
